import pandas as pd
import time
from openpyxl import load_workbook


def clear_excel_file(file_path):
    try:
        # Excel dosyasını yükleyin.
        wb = load_workbook(file_path)

        # İşlem yapmak istediğiniz çalışma sayfasını (worksheet) seçin.
        ws = wb.active  # Aktif çalışma sayfasını seçiyoruz, isterseniz farklı bir sayfayı da seçebilirsiniz.

        # Verileri silme (tüm verileri siliyoruz).
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # Dosyayı kaydedin.
        wb.save(file_path)
        wb.close()

        print("Excel dosyasındaki veriler silindi.")

    except Exception as e:
        print(f"Hata: {str(e)}")

def append_to_excel():
    # Verilerinizi mevcut Excel dosyasına eklemek için dosyayı açın (var olan bir dosyaysa).
    try:
        df = pd.read_excel("veriler.xlsx")
    except FileNotFoundError:
        # Eğer dosya bulunamazsa yeni bir dosya oluşturun.
        df = pd.DataFrame(columns=["Class", "Count", "Time"])

    while True:
        a_classname = "Class"
        a_classvalue = "Person"

        a_vername = "Count"
        a_vervalue = 1

        timer = time.strftime("%H:%M:%S")
        timename = "Time"

        new_data = {
            a_classname: [a_classvalue],
            a_vername: [a_vervalue],
            timename: [timer]
        }

        # Yeni veriyi DataFrame'e ekleyin.
        new_row = pd.DataFrame(new_data)
        df = pd.concat([df, new_row], ignore_index=True)

        # Veriyi Excel dosyasına kaydedin.
        df.to_excel("veriler.xlsx", index=False)

        # Her saniye bekleme yapın.
        time.sleep(1)

if __name__ == "__main__":
    excel_file_path = "veriler.xlsx"  # Sileceğiniz Excel dosyasının yolunu belirtin.
    clear_excel_file(excel_file_path)
    append_to_excel()
